"""Phase 4 (Layer 4) — coverage + precision verification vs parts_table ground truth.

This is the LAYER 4 accuracy check from docs/07-output-verification.md, adapted
for the actual data structure: one sheet has ONE parts table but MULTIPLE
schematic images, each showing a SUBSET of callouts.

Two complementary metrics:

  PER-IMAGE PRECISION
    precision = |OCR ∩ parts_table_sheet| / |OCR|
    Answers: "Of the callouts OCR found, how many are real (vs hallucinated)?"
    Catches: hallucinated callout numbers (model invented n=99 that doesn't exist)
    Cannot catch: missing callouts (OCR didn't return n=5 that's on the image)

  PER-SHEET UNION COVERAGE
    coverage = |union(OCR across all images for sheet) ∩ parts_table| / |parts_table|
    Answers: "Across all images for this sheet, are all parts callout-mapped?"
    Catches: parts that no image's OCR found (system-level miss)

Confidence re-tiering policy:
  - Use PRECISION for per-image confidence demotion (high precision → keep tier,
    low precision → demote because hallucination is a real defect)
  - Do NOT use precision to UPGRADE — Phase 3a Layer 3 warnings still apply
  - Annotate each record with precision + parent sheet union_coverage for
    downstream consumers (e.g. marketplace listing pipeline can decide)

Input:
  - shared/mlx-vision-output-final.jsonl  (from phase3_verify)
  - shared/sample-data/example.xlsx       (ground truth parts tables)
  - shared/vision-mlx/extracted_images/manifest.csv (image→sheet mapping)

Output:
  - shared/mlx-vision-output-final-with-coverage.jsonl
  - Console summary: precision distribution + union coverage per sheet
"""
from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[2]
SHARED_DIR = REPO_ROOT / "shared"
XLSX_PATH = SHARED_DIR / "sample-data" / "example.xlsx"
MANIFEST_CSV = Path(__file__).resolve().parent / "extracted_images" / "manifest.csv"
INPUT_FILE = SHARED_DIR / "mlx-vision-output-final.jsonl"
OUTPUT_FILE = SHARED_DIR / "mlx-vision-output-final-with-coverage.jsonl"


def extract_callouts_from_xlsx() -> dict[str, set[int]]:
    """Per sheet → set of callout numbers from parts table.

    Excel stores integers as floats (1.0, 2.0) by default. Heuristic:
    scan first 6 columns, accept ints, exact-int-valued floats, and
    string-encoded numbers ("1.", "(1)").
    """
    print(f"[phase4] Loading xlsx (~242 MB, takes ~30s)...")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    out: dict[str, set[int]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        col_candidates: list[tuple[int, list[int]]] = []
        for col_idx in range(1, 7):
            ints_in_col: list[int] = []
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
                v = row[0]
                n: int | None = None
                if isinstance(v, int) and 0 < v < 1000:
                    n = v
                elif isinstance(v, float) and 0 < v < 1000 and v == int(v):
                    n = int(v)
                elif isinstance(v, str):
                    m = re.search(r"\b(\d{1,3})\b", v)
                    if m:
                        candidate = int(m.group(1))
                        if 0 < candidate < 1000:
                            n = candidate
                if n is not None:
                    ints_in_col.append(n)
            if len(ints_in_col) >= 3:
                col_candidates.append((col_idx, ints_in_col))

        if not col_candidates:
            out[sheet_name] = set()
            continue

        # Pick column with most callout-like ints
        best = max(col_candidates, key=lambda c: len(c[1]))
        out[sheet_name] = set(best[1])

    wb.close()
    n_with_callouts = sum(1 for v in out.values() if v)
    print(f"[phase4] Extracted callouts for {n_with_callouts}/{len(out)} sheets")
    return out


def load_image_sheets() -> dict[str, list[str]]:
    """sha256_short (12 chars) → list of sheet names."""
    mapping: dict[str, set[str]] = defaultdict(set)
    with MANIFEST_CSV.open() as f:
        for row in csv.DictReader(f):
            sha_short = row["sha256"][:12]
            mapping[sha_short].add(row["sheet_name"])
    return {sha: sorted(sheets) for sha, sheets in mapping.items()}


def compute_per_image_precision(
    ocr_n: set[int], parts_n_for_sheets: set[int]
) -> tuple[float | None, dict]:
    """Precision = |OCR ∩ parts| / |OCR|.

    None precision = ground truth unavailable for this image's sheet(s).
    """
    if not parts_n_for_sheets:
        return None, {"reason": "no_parts_table_for_sheet"}
    if not ocr_n:
        return None, {"reason": "ocr_empty"}

    matched = ocr_n & parts_n_for_sheets
    extra = ocr_n - parts_n_for_sheets
    precision = len(matched) / len(ocr_n)

    return precision, {
        "precision": round(precision, 3),
        "ocr_count": len(ocr_n),
        "matched_in_parts_table": len(matched),
        "hallucinated_callouts": len(extra),
        "extra_examples": sorted(extra)[:5],
    }


def adjust_confidence(
    prior_confidence: str,
    prior_warnings: list[str],
    precision: float | None,
) -> tuple[str, str]:
    """Apply precision-based adjustment to Phase 3a confidence.

    Policy:
      - DEAD stays DEAD (no OCR data to verify).
      - precision ≥ 0.9    → keep prior tier (no upgrade because recall unmeasured)
      - precision 0.7-0.9  → demote one tier if currently HIGH
      - precision < 0.7    → demote to LOW (definite hallucination present)
      - precision None     → keep prior (no ground truth for this image)
    """
    if prior_confidence == "dead":
        return "dead", "DEAD records unchanged"

    if precision is None:
        return prior_confidence, "no_ground_truth_keep_prior"

    if precision >= 0.9:
        return prior_confidence, "precision_high_keep_prior"

    if precision >= 0.7:
        if prior_confidence == "high":
            return "medium", "precision_medium_demoted_from_high"
        return prior_confidence, "precision_medium_keep_prior"

    # precision < 0.7 → significant hallucination
    return "low", "precision_low_demoted"


def main() -> int:
    if not INPUT_FILE.exists():
        print(f"[phase4] ERROR: {INPUT_FILE} not found. Run phase3_verify.py first.")
        return 1

    parts_by_sheet = extract_callouts_from_xlsx()
    image_to_sheets = load_image_sheets()

    records: list[dict] = []
    for line in INPUT_FILE.read_text().splitlines():
        records.append(json.loads(line))
    print(f"[phase4] Loaded {len(records)} records from {INPUT_FILE.name}")

    # Build sha → record map AND sheet → list of OCR sets (for union coverage)
    sheet_ocr_union: dict[str, set[int]] = defaultdict(set)
    for rec in records:
        sha = rec["sha256"]
        sheets = image_to_sheets.get(sha, [])
        ocr_list = rec.get("ocr_result") or []
        ocr_n: set[int] = set()
        if isinstance(ocr_list, list):
            for c in ocr_list:
                if isinstance(c, dict) and isinstance(c.get("n"), int):
                    ocr_n.add(c["n"])
        for s in sheets:
            sheet_ocr_union[s] |= ocr_n

    # Per-image precision
    tier_before = Counter()
    tier_after = Counter()
    precision_buckets = Counter()
    no_ground_truth = 0

    for rec in records:
        sha = rec["sha256"]
        tier_before[rec.get("confidence", "low")] += 1

        sheets = image_to_sheets.get(sha, [])
        parts_n: set[int] = set()
        for s in sheets:
            parts_n |= parts_by_sheet.get(s, set())

        ocr_list = rec.get("ocr_result") or []
        ocr_n: set[int] = set()
        if isinstance(ocr_list, list):
            for c in ocr_list:
                if isinstance(c, dict) and isinstance(c.get("n"), int):
                    ocr_n.add(c["n"])

        precision, detail = compute_per_image_precision(ocr_n, parts_n)

        if precision is None:
            no_ground_truth += 1
            precision_buckets["no_truth"] += 1
        elif precision >= 0.9:
            precision_buckets["≥90%"] += 1
        elif precision >= 0.7:
            precision_buckets["70-90%"] += 1
        elif precision > 0:
            precision_buckets["0-70%"] += 1
        else:
            precision_buckets["0%"] += 1

        new_conf, reason = adjust_confidence(
            rec.get("confidence", "low"),
            rec.get("warnings", []),
            precision,
        )

        rec["precision"] = precision
        rec["precision_detail"] = detail
        rec["confidence_layer3"] = rec.get("confidence")
        rec["confidence"] = new_conf
        rec["confidence_change_reason"] = reason
        tier_after[new_conf] += 1

    # Per-sheet union coverage
    sheet_coverage_stats = []
    for sheet, parts_n in parts_by_sheet.items():
        if not parts_n:
            continue
        union_ocr = sheet_ocr_union.get(sheet, set())
        matched = union_ocr & parts_n
        coverage = len(matched) / len(parts_n) if parts_n else 0
        sheet_coverage_stats.append({
            "sheet": sheet,
            "parts_count": len(parts_n),
            "union_ocr_matched": len(matched),
            "coverage": coverage,
        })

    sheet_coverage_stats.sort(key=lambda x: x["coverage"])
    full_cov_sheets = sum(1 for s in sheet_coverage_stats if s["coverage"] >= 1.0)
    high_cov = sum(1 for s in sheet_coverage_stats if s["coverage"] >= 0.7)

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT_FILE.open("w") as out:
        for rec in records:
            out.write(json.dumps(rec, ensure_ascii=False) + "\n")

    print()
    print(f"[phase4] ============= SUMMARY =============")
    print(f"[phase4] Records processed: {len(records)}")
    print(f"[phase4] Records with no ground truth: {no_ground_truth}")
    print(f"[phase4]")
    print(f"[phase4] Per-image PRECISION (catches hallucinated callouts):")
    for bucket in ("≥90%", "70-90%", "0-70%", "0%", "no_truth"):
        n = precision_buckets[bucket]
        print(f"[phase4]   {bucket:10s}: {n:>4} ({100*n/len(records):.1f}%)")
    print(f"[phase4]")
    print(f"[phase4] Confidence tier — BEFORE (Phase 3a) vs AFTER (Phase 4 precision):")
    print(f"[phase4]   {'tier':10s} {'before':>8} {'after':>8}  Δ")
    for tier in ("high", "medium", "low", "dead"):
        b = tier_before[tier]
        a = tier_after[tier]
        delta = a - b
        sign = "+" if delta >= 0 else ""
        print(f"[phase4]   {tier:10s} {b:>8} {a:>8}  {sign}{delta}")
    print(f"[phase4]")
    print(f"[phase4] Per-sheet UNION COVERAGE (across all images per sheet):")
    print(f"[phase4]   Sheets with ground truth: {len(sheet_coverage_stats)}")
    print(f"[phase4]   100% union coverage:      {full_cov_sheets} ({100*full_cov_sheets/len(sheet_coverage_stats):.1f}%)")
    print(f"[phase4]   ≥70% union coverage:      {high_cov} ({100*high_cov/len(sheet_coverage_stats):.1f}%)")
    print(f"[phase4]")
    print(f"[phase4] Lowest 5 sheet coverages:")
    for s in sheet_coverage_stats[:5]:
        print(f"[phase4]   {s['sheet'][:35]:35s}  parts={s['parts_count']:>3}  matched={s['union_ocr_matched']:>3}  cov={s['coverage']:.1%}")
    print(f"[phase4]")
    print(f"[phase4] Highest 5 sheet coverages:")
    for s in sheet_coverage_stats[-5:]:
        print(f"[phase4]   {s['sheet'][:35]:35s}  parts={s['parts_count']:>3}  matched={s['union_ocr_matched']:>3}  cov={s['coverage']:.1%}")
    print(f"[phase4]")
    print(f"[phase4] Output: {OUTPUT_FILE}")
    print(f"[phase4] DONE")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
