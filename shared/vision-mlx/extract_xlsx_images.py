"""Extract embedded images from the Kayo catalog xlsx.

Walks every sheet, pulls each embedded image, writes it to disk under
`out_dir/<sheet_name>/<index>.<ext>` and emits a manifest CSV mapping image
path -> sheet name -> anchor cell coordinates.

The manifest is what the downstream pipeline joins against parsed tabular rows
to figure out which schematic belongs to which model section.
"""
from __future__ import annotations

import csv
import hashlib
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


SAFE_NAME_RE = re.compile(r"[^A-Za-z0-9_-]+")


def safe_name(s: str) -> str:
    return SAFE_NAME_RE.sub("_", s)[:80] or "unnamed"


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def extract(xlsx_path: Path, out_dir: Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    manifest_path = out_dir / "manifest.csv"

    print(f"[extract] opening {xlsx_path.name} ...")
    wb = load_workbook(xlsx_path, data_only=True)
    sheets = wb.sheetnames
    print(f"[extract] {len(sheets)} sheets")

    total_images = 0
    seen_hashes: dict[str, str] = {}

    with manifest_path.open("w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([
            "sheet_name", "image_index", "sha256", "file_path", "ext",
            "size_bytes", "anchor_cell", "dedup_of",
        ])

        for sheet_idx, sheet_name in enumerate(sheets):
            ws = wb[sheet_name]
            images = getattr(ws, "_images", []) or []
            if not images:
                continue
            sheet_dir = out_dir / f"{sheet_idx:03d}_{safe_name(sheet_name)}"
            sheet_dir.mkdir(parents=True, exist_ok=True)

            for img_idx, img in enumerate(images):
                # openpyxl Image objects have _data() callable returning bytes
                try:
                    data = img._data() if callable(getattr(img, "_data", None)) else img.ref
                except Exception as e:
                    print(f"[extract]  ! {sheet_name}#{img_idx} read fail: {e}")
                    continue
                if not data:
                    continue

                ext = getattr(img, "format", "png") or "png"
                ext = ext.lower().lstrip(".")
                h = sha256_bytes(data)
                file_path = sheet_dir / f"{img_idx:03d}_{h[:12]}.{ext}"

                dedup_of = seen_hashes.get(h)
                if dedup_of is None:
                    file_path.write_bytes(data)
                    seen_hashes[h] = str(file_path)
                else:
                    # Already wrote this exact image elsewhere; skip writing
                    file_path = Path(dedup_of)

                anchor = ""
                try:
                    a = img.anchor._from
                    anchor = f"{a.col}:{a.row}"
                except Exception:
                    pass

                writer.writerow([
                    sheet_name, img_idx, h,
                    str(file_path.relative_to(out_dir)),
                    ext, len(data), anchor,
                    dedup_of if dedup_of else "",
                ])
                total_images += 1

            print(f"[extract] {sheet_name}: {len(images)} images")

    print(f"\n[extract] total: {total_images} images, "
          f"{len(seen_hashes)} unique, manifest at {manifest_path}")


def main() -> int:
    if len(sys.argv) < 2:
        print("Usage: python extract_xlsx_images.py <xlsx_path> [out_dir]")
        return 1
    xlsx_path = Path(sys.argv[1]).resolve()
    out_dir = Path(sys.argv[2] if len(sys.argv) > 2 else "extracted_images").resolve()
    if not xlsx_path.exists():
        print(f"ERROR: {xlsx_path} not found")
        return 2
    extract(xlsx_path, out_dir)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
