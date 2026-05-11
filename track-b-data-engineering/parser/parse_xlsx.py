"""Top-level xlsx parser glue.

Reads every sheet in the source xlsx, detects sections, normalises rows,
attaches per-sheet fitment, and yields product records ready to write to
either CSV (for Track A parity comparison) or Iceberg (for Dagster path).

This is the integration layer that ties the four parser modules together
so Track B's silver layer can produce semantically equivalent output to
Track A given the same input xlsx.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook

from .fitment_resolver import ResolvedFitment, resolve_fitment_from_sheet_name
from .row_normalizer import NormalisedRow, _Reject, normalise_row
from .section_detector import SheetRow, detect_sections

log = logging.getLogger(__name__)


@dataclass(frozen=True)
class FitmentEntry:
    """Single fitment record attached to a part.

    Wire-format matches Track A's `products.fitment` JSONB array element:
    {make, year, model, model_code, variant, section, callout_no, confidence}
    """

    make: str
    year: int | None
    model: str | None
    model_code: str | None
    variant: str | None
    section: str | None
    callout_no: str | None
    confidence: str  # "high" | "medium" | "low"


@dataclass(frozen=True)
class ParsedProduct:
    """One product row, ready for persistence in either track's storage."""

    part_number: str
    name_en: str | None
    name_cn: str | None
    spec_cn: str | None
    retail_price: float | None
    fitment: list[FitmentEntry]


def parse_xlsx(path: Path | str, *, make: str = "Kayo") -> list[ParsedProduct]:
    """Parse the OEM xlsx end-to-end.

    Returns one ParsedProduct per (sheet, part_number) — the same shape
    Track A persists into PostgreSQL.products. Multiple rows for the same
    part_number across sheets remain as separate ParsedProduct entries;
    deduplication happens in the silver merge step (PG ON CONFLICT or
    Iceberg MERGE INTO).
    """
    products: list[ParsedProduct] = []
    for sheet_name, rows in _iter_sheets(path):
        # fitment may be None for sheets like "CARBURETOR JETS" / "SPARK PLUGS"
        # — Track A still extracts the rows in these sheets, just with an
        # empty fitment array. Mirror that here.
        fitment = resolve_fitment_from_sheet_name(sheet_name)

        sections = detect_sections(rows)
        if not sections:
            log.debug("No sections detected in sheet: %s", sheet_name)
            continue

        for section in sections:
            for row in _rows_in_section(rows, section):
                result = normalise_row(row, section, sheet_name)
                if isinstance(result, _Reject):
                    continue
                product = _to_product(result, fitment, section.title, make)
                products.append(product)

    return products


def _iter_sheets(path: Path | str) -> Iterator[tuple[str, list[SheetRow]]]:
    """Yield (sheet_name, rows) for every visible sheet.

    Uses openpyxl read-only mode to keep RAM bounded — the 230 MB xlsx
    has 1,586 embedded images that we don't need at this layer.
    """
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            rows: list[SheetRow] = []
            for row_idx, raw_row in enumerate(ws.iter_rows(values_only=True), start=1):
                # Align to 1-based by prepending None at index 0 — matches
                # the (off-by-one safe) shape Track A's xlsx-reader emits.
                values = (None, *raw_row)
                rows.append(SheetRow(row_index=row_idx, values=values))
            yield ws.title, rows
    finally:
        wb.close()


def _rows_in_section(rows: list[SheetRow], section) -> list[SheetRow]:  # type: ignore[no-untyped-def]
    return [
        r for r in rows
        if section.data_start_index <= r.row_index <= section.data_end_index
    ]


def _to_product(
    row: NormalisedRow,
    fitment: ResolvedFitment | None,
    section_title: str | None,
    make: str,
) -> ParsedProduct:
    """Build the product record.

    `fitment=None` means the source sheet is a reference/spec sheet
    (carburetor jets, fork seal specs, etc.) — Track A still emits the
    product row but with an empty fitment array. Mirror that.
    """
    if fitment is None:
        return ParsedProduct(
            part_number=row.part_number,
            name_en=row.name_en,
            name_cn=row.name_cn,
            spec_cn=row.spec_cn,
            retail_price=row.retail_price,
            fitment=[],
        )

    # Year resolution: emit a single fitment with year=year_start.
    # Track A's products.fitment array carries one entry per (part, sheet)
    # tuple. When the sheet name carries no year info, Track A coalesces
    # to 0 (the JSONB default), not null — mirror that wire format.
    callout_raw = row.callout.raw if row.callout else None
    year = fitment.year_start if fitment.year_start is not None else 0
    fitments = [
        FitmentEntry(
            make=make,
            year=year,
            model=fitment.model_code,
            model_code=fitment.model_code,
            variant=fitment.variant,
            section=section_title,
            callout_no=callout_raw,
            confidence="high" if fitment.model_code else "medium",
        )
    ]

    return ParsedProduct(
        part_number=row.part_number,
        name_en=row.name_en,
        name_cn=row.name_cn,
        spec_cn=row.spec_cn,
        retail_price=row.retail_price,
        fitment=fitments,
    )


def _year_range(start: int | None, end: int | None) -> list[int | None]:
    """Expand a year window to a list of explicit years.

    None when both endpoints are None — caller emits one fitment with year=None.
    Single year when start==end. Otherwise inclusive range.
    Open-ended (start, None) emits just [start].
    """
    if start is None and end is None:
        return [None]
    if start is not None and end is None:
        return [start]
    if start is not None and end is not None:
        if start == end:
            return [start]
        return list(range(start, end + 1))
    return [None]
